import os
import pandas as pd
import numpy as np
import math
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

source_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']

source_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

target_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

root_path = 'D:/ScottTsai/res_multi/auto_hyperparameter/'
layer_path = 'imp_v3_add3_with_kmm_dr_5Layer_auto_all/'
result_path = 'accuracy/result'
excel_path = 'accuracy/result/'

goal_excel = load_workbook(root_path + layer_path + excel_path + "all_to_all.xlsx")
if '1_to_2_average' not in goal_excel.sheetnames:
    create_excel_sheet = goal_excel.create_sheet("1_to_2_average")

all_result_sheet = goal_excel.get_sheet_by_name('All_accuracy')



target_com_list = [12, 13, 14, 15, 16, 17, 18,
                   23, 24, 25, 26, 27, 28,
                   34, 35, 36, 37, 38,
                   45, 46,47, 48,
                   56, 57, 58,
                   67, 68,
                   78]
ws = goal_excel.get_sheet_by_name("1_to_2_average")
goal_excel.active = ws

goal_excel.active.merge_cells('C2:BF2')
goal_excel.active.merge_cells('C15:BF15')
goal_excel.active.merge_cells('C28:BF28')
goal_excel.active.merge_cells('C41:BF41')
goal_excel.active.merge_cells('C54:BF54')
goal_excel.active.merge_cells('C67:BF67')
goal_excel.active.merge_cells('C80:BF80')
goal_excel.active.merge_cells('C93:BF93')

goal_excel.active.column_dimensions['B'].width = 22.29
ws['C2'] = 1
ws['C2'].alignment = Alignment(horizontal= 'center')
ws['B2'] = 'source'
ws['B2'].alignment = Alignment(horizontal= 'center')
ws['B3'] = 'target'
ws['B3'].alignment = Alignment(horizontal= 'center')
ws['B4'] = 'Tsai'
ws['B4'].alignment = Alignment(horizontal= 'center')
ws['B5'] = 'Acc_positive'
ws['B5'].alignment = Alignment(horizontal= 'center')
ws['B6'] = 'Acc_total'
ws['B6'].alignment = Alignment(horizontal= 'center')

ws['C15'] = 2
ws['C15'].alignment = Alignment(horizontal= 'center')
ws['B15'] = 'source'
ws['B15'].alignment = Alignment(horizontal= 'center')
ws['B16'] = 'target'
ws['B16'].alignment = Alignment(horizontal= 'center')
ws['B17'] = 'Tsai'
ws['B17'].alignment = Alignment(horizontal= 'center')
ws['B18'] = 'Acc_positive'
ws['B18'].alignment = Alignment(horizontal= 'center')
ws['B19'] = 'Acc_total'
ws['B19'].alignment = Alignment(horizontal= 'center')

ws['C28'] = 3
ws['C28'].alignment = Alignment(horizontal= 'center')
ws['B28'] = 'source'
ws['B28'].alignment = Alignment(horizontal= 'center')
ws['B29'] = 'target'
ws['B29'].alignment = Alignment(horizontal= 'center')
ws['B30'] = 'Tsai'
ws['B30'].alignment = Alignment(horizontal= 'center')
ws['B31'] = 'Acc_positive'
ws['B31'].alignment = Alignment(horizontal= 'center')
ws['B32'] = 'Acc_total'
ws['B32'].alignment = Alignment(horizontal= 'center')


ws['C41'] = 4
ws['C41'].alignment = Alignment(horizontal= 'center')
ws['B41'] = 'source'
ws['B41'].alignment = Alignment(horizontal= 'center')
ws['B42'] = 'target'
ws['B42'].alignment = Alignment(horizontal= 'center')
ws['B43'] = 'Tsai'
ws['B43'].alignment = Alignment(horizontal= 'center')
ws['B44'] = 'Acc_positive'
ws['B44'].alignment = Alignment(horizontal= 'center')
ws['B45'] = 'Acc_total'
ws['B45'].alignment = Alignment(horizontal= 'center')

ws['C54'] = 5
ws['C54'].alignment = Alignment(horizontal= 'center')
ws['B54'] = 'source'
ws['B54'].alignment = Alignment(horizontal= 'center')
ws['B55'] = 'target'
ws['B55'].alignment = Alignment(horizontal= 'center')
ws['B56'] = 'Tsai'
ws['B56'].alignment = Alignment(horizontal= 'center')
ws['B57'] = 'Acc_positive'
ws['B57'].alignment = Alignment(horizontal= 'center')
ws['B58'] = 'Acc_total'
ws['B58'].alignment = Alignment(horizontal= 'center')

ws['C67'] = 6
ws['C67'].alignment = Alignment(horizontal= 'center')
ws['B67'] = 'source'
ws['B67'].alignment = Alignment(horizontal= 'center')
ws['B68'] = 'target'
ws['B68'].alignment = Alignment(horizontal= 'center')
ws['B69'] = 'Tsai'
ws['B69'].alignment = Alignment(horizontal= 'center')
ws['B70'] = 'Acc_positive'
ws['B70'].alignment = Alignment(horizontal= 'center')
ws['B71'] = 'Acc_total'
ws['B71'].alignment = Alignment(horizontal= 'center')

ws['C80'] = 7
ws['C80'].alignment = Alignment(horizontal= 'center')
ws['B80'] = 'source'
ws['B80'].alignment = Alignment(horizontal= 'center')
ws['B81'] = 'target'
ws['B81'].alignment = Alignment(horizontal= 'center')
ws['B82'] = 'Tsai'
ws['B82'].alignment = Alignment(horizontal= 'center')
ws['B83'] = 'Acc_positive'
ws['B83'].alignment = Alignment(horizontal= 'center')
ws['B84'] = 'Acc_total'
ws['B84'].alignment = Alignment(horizontal= 'center')

ws['C93'] = 8
ws['C93'].alignment = Alignment(horizontal= 'center')
ws['B93'] = 'source'
ws['B93'].alignment = Alignment(horizontal= 'center')
ws['B94'] = 'target'
ws['B94'].alignment = Alignment(horizontal= 'center')
ws['B95'] = 'Tsai'
ws['B95'].alignment = Alignment(horizontal= 'center')
ws['B96'] = 'Acc_positive'
ws['B96'].alignment = Alignment(horizontal= 'center')
ws['B97'] = 'Acc_total'
ws['B97'].alignment = Alignment(horizontal= 'center')

count = 0
for i in range (3, 59, 2):
    ws.merge_cells(start_row=3, start_column=i, end_row=3, end_column=i+1)
    ws.cell(row=3, column=i).value = target_com_list[count]
    ws.cell(row=3, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=i).value = '平均'
    ws.cell(row=4, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=i + 1).value = '正負'
    ws.cell(row=4, column=i + 1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=16, start_column=i, end_row=16, end_column=i+1)
    ws.cell(row=16, column=i).value = target_com_list[count]
    ws.cell(row=16, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=17, column=i).value = '平均'
    ws.cell(row=17, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=17, column=i + 1).value = '正負'
    ws.cell(row=17, column=i + 1).alignment = Alignment(horizontal='center')


    ws.merge_cells(start_row=29, start_column=i, end_row=29, end_column=i+1)
    ws.cell(row=29, column=i).value = target_com_list[count]
    ws.cell(row=29, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=30, column=i).value = '平均'
    ws.cell(row=30, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=30, column=i + 1).value = '正負'
    ws.cell(row=30, column=i + 1).alignment = Alignment(horizontal='center')


    ws.merge_cells(start_row=42, start_column=i, end_row=42, end_column=i+1)
    ws.cell(row=42, column=i).value = target_com_list[count]
    ws.cell(row=42, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=43, column=i).value = '平均'
    ws.cell(row=43, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=43, column=i + 1).value = '正負'
    ws.cell(row=43, column=i + 1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=55, start_column=i, end_row=55, end_column=i+1)
    ws.cell(row=55, column=i).value = target_com_list[count]
    ws.cell(row=55, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=56, column=i).value = '平均'
    ws.cell(row=56, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=56, column=i + 1).value = '正負'
    ws.cell(row=56, column=i + 1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=68, start_column=i, end_row=68, end_column=i+1)
    ws.cell(row=68, column=i).value = target_com_list[count]
    ws.cell(row=68, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=69, column=i).value = '平均'
    ws.cell(row=69, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=69, column=i + 1).value = '正負'
    ws.cell(row=69, column=i + 1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=81, start_column=i, end_row=81, end_column=i+1)
    ws.cell(row=81, column=i).value = target_com_list[count]
    ws.cell(row=81, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=82, column=i).value = '平均'
    ws.cell(row=82, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=82, column=i + 1).value = '正負'
    ws.cell(row=82, column=i + 1).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=94, start_column=i, end_row=94, end_column=i+1)
    ws.cell(row=94, column=i).value = target_com_list[count]
    ws.cell(row=94, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=95, column=i).value = '平均'
    ws.cell(row=95, column=i).alignment = Alignment(horizontal='center')
    ws.cell(row=95, column=i + 1).value = '正負'
    ws.cell(row=95, column=i + 1).alignment = Alignment(horizontal='center')

    count = count + 1

#one-to-two average

for row_index in range(5, 108, 13):
    cell_column_count = 3
    for index_1 in range (3, 19, 2):
        for index_2 in range ((index_1 + 2), 19, 2):
            value_1_pos = all_result_sheet.cell(row=row_index, column=index_1).value
            value_2_pos = all_result_sheet.cell(row=row_index, column=index_2).value
            diff_1_pos = all_result_sheet.cell(row=row_index, column = (index_1 + 1)).value
            diff_2_pos = all_result_sheet.cell(row=row_index, column = (index_2 + 1)).value

            avg_pos = (value_1_pos + value_2_pos) / 2
            avg_diff_pos = (diff_1_pos + diff_2_pos) / 2
            avg_pos = np.round(avg_pos, 2)
            avg_diff_pos = np.round(avg_diff_pos, 2)

            ws.cell(row=row_index, column=cell_column_count).value = avg_pos
            ws.cell(row=row_index, column=cell_column_count).alignment = Alignment(horizontal='center')
            ws.cell(row=row_index, column=(cell_column_count + 1)).value = avg_diff_pos
            ws.cell(row=row_index, column=(cell_column_count + 1)).alignment = Alignment(horizontal='center')

            value_1_total = all_result_sheet.cell(row=(row_index + 1), column=index_1).value
            value_2_total = all_result_sheet.cell(row=(row_index + 1), column=index_2).value
            diff_1_total = all_result_sheet.cell(row=(row_index + 1), column = (index_2 + 1)).value
            diff_2_total = all_result_sheet.cell(row=(row_index + 1), column = (index_2 + 1)).value
            avg_total = (value_1_total + value_2_total) / 2
            avg_diff_total = (diff_1_total + diff_2_total) / 2
            avg_total = np.round(avg_total, 2)
            avg_diff_total = np.round(avg_diff_total, 2)

            ws.cell(row=(row_index + 1), column=cell_column_count).value = avg_total
            ws.cell(row=(row_index + 1), column=cell_column_count).alignment = Alignment(horizontal='center')
            ws.cell(row=(row_index + 1), column=(cell_column_count + 1)).value = avg_diff_total
            ws.cell(row=(row_index + 1), column=(cell_column_count + 1)).alignment = Alignment(horizontal='center')



            cell_column_count = cell_column_count +2










goal_excel.save(root_path + layer_path + excel_path + "all_to_all.xlsx")


