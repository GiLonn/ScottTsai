import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


source_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']

source_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

target_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}
'''
target_com_list = [12, 13, 14, 15, 16, 17, 18,
                   23, 24, 25, 26, 27, 28,
                   34, 35, 36, 37, 38,
                   45, 46, 47, 48,
                   56, 57, 58,
                   67, 68,
                   78]
'''
root_path = 'D:/ScottTsai/res_multi/auto_hyperparameter/'
layer_path = 'imp_v3_add3_with_kmm_dr_5Layer_auto_all/'
result_path = 'accuracy/result'
excel_path = 'accuracy/result/'

wb = Workbook()
std = wb.get_sheet_by_name('Sheet')
std.title = 'All_accuracy'
wb.active.merge_cells('C2:R2')
wb.active.merge_cells('C15:R15')
wb.active.merge_cells('C28:R28')
wb.active.merge_cells('C41:R41')
wb.active.merge_cells('C54:R54')
wb.active.merge_cells('C67:R67')
wb.active.merge_cells('C80:R80')
wb.active.merge_cells('C93:R93')

ws = wb.get_sheet_by_name('All_accuracy')
wb.active.column_dimensions['B'].width = 22.29

ws['C2'] = 1
ws['C2'].alignment = Alignment(horizontal= 'center')
ws['B2'] = 'source'
ws['B2'].alignment = Alignment(horizontal= 'center')
ws['B3'] = 'target'
ws['B3'].alignment = Alignment(horizontal= 'center')
ws['B4'] = 'Tsai'
ws['B4'].alignment = Alignment(horizontal= 'center')
ws['B5'] = 'Acc_positive'
ws['B5'].alignment = Alignment(horizontal= 'center')
ws['B6'] = 'Acc_total'
ws['B6'].alignment = Alignment(horizontal= 'center')

ws['C15'] = 2
ws['C15'].alignment = Alignment(horizontal= 'center')
ws['B15'] = 'source'
ws['B15'].alignment = Alignment(horizontal= 'center')
ws['B16'] = 'target'
ws['B16'].alignment = Alignment(horizontal= 'center')
ws['B17'] = 'Tsai'
ws['B17'].alignment = Alignment(horizontal= 'center')
ws['B18'] = 'Acc_positive'
ws['B18'].alignment = Alignment(horizontal= 'center')
ws['B19'] = 'Acc_total'
ws['B19'].alignment = Alignment(horizontal= 'center')

ws['C28'] = 3
ws['C28'].alignment = Alignment(horizontal= 'center')
ws['B28'] = 'source'
ws['B28'].alignment = Alignment(horizontal= 'center')
ws['B29'] = 'target'
ws['B29'].alignment = Alignment(horizontal= 'center')
ws['B30'] = 'Tsai'
ws['B30'].alignment = Alignment(horizontal= 'center')
ws['B31'] = 'Acc_positive'
ws['B31'].alignment = Alignment(horizontal= 'center')
ws['B32'] = 'Acc_total'
ws['B32'].alignment = Alignment(horizontal= 'center')


ws['C41'] = 4
ws['C41'].alignment = Alignment(horizontal= 'center')
ws['B41'] = 'source'
ws['B41'].alignment = Alignment(horizontal= 'center')
ws['B42'] = 'target'
ws['B42'].alignment = Alignment(horizontal= 'center')
ws['B43'] = 'Tsai'
ws['B43'].alignment = Alignment(horizontal= 'center')
ws['B44'] = 'Acc_positive'
ws['B44'].alignment = Alignment(horizontal= 'center')
ws['B45'] = 'Acc_total'
ws['B45'].alignment = Alignment(horizontal= 'center')

ws['C54'] = 5
ws['C54'].alignment = Alignment(horizontal= 'center')
ws['B54'] = 'source'
ws['B54'].alignment = Alignment(horizontal= 'center')
ws['B55'] = 'target'
ws['B55'].alignment = Alignment(horizontal= 'center')
ws['B56'] = 'Tsai'
ws['B56'].alignment = Alignment(horizontal= 'center')
ws['B57'] = 'Acc_positive'
ws['B57'].alignment = Alignment(horizontal= 'center')
ws['B58'] = 'Acc_total'
ws['B58'].alignment = Alignment(horizontal= 'center')

ws['C67'] = 6
ws['C67'].alignment = Alignment(horizontal= 'center')
ws['B67'] = 'source'
ws['B67'].alignment = Alignment(horizontal= 'center')
ws['B68'] = 'target'
ws['B68'].alignment = Alignment(horizontal= 'center')
ws['B69'] = 'Tsai'
ws['B69'].alignment = Alignment(horizontal= 'center')
ws['B70'] = 'Acc_positive'
ws['B70'].alignment = Alignment(horizontal= 'center')
ws['B71'] = 'Acc_total'
ws['B71'].alignment = Alignment(horizontal= 'center')

ws['C80'] = 7
ws['C80'].alignment = Alignment(horizontal= 'center')
ws['B80'] = 'source'
ws['B80'].alignment = Alignment(horizontal= 'center')
ws['B81'] = 'target'
ws['B81'].alignment = Alignment(horizontal= 'center')
ws['B82'] = 'Tsai'
ws['B82'].alignment = Alignment(horizontal= 'center')
ws['B83'] = 'Acc_positive'
ws['B83'].alignment = Alignment(horizontal= 'center')
ws['B84'] = 'Acc_total'
ws['B84'].alignment = Alignment(horizontal= 'center')

ws['C93'] = 8
ws['C93'].alignment = Alignment(horizontal= 'center')
ws['B93'] = 'source'
ws['B93'].alignment = Alignment(horizontal= 'center')
ws['B94'] = 'target'
ws['B94'].alignment = Alignment(horizontal= 'center')
ws['B95'] = 'Tsai'
ws['B95'].alignment = Alignment(horizontal= 'center')
ws['B96'] = 'Acc_positive'
ws['B96'].alignment = Alignment(horizontal= 'center')
ws['B97'] = 'Acc_total'
ws['B97'].alignment = Alignment(horizontal= 'center')

for i in range (3, 19, 2):
    ws.merge_cells(start_row=3, start_column=i, end_row=3, end_column=i+1)
    ws.merge_cells(start_row=16, start_column=i, end_row=16, end_column=i+1)
    ws.merge_cells(start_row=29, start_column=i, end_row=29, end_column=i+1)
    ws.merge_cells(start_row=42, start_column=i, end_row=42, end_column=i+1)
    ws.merge_cells(start_row=55, start_column=i, end_row=55, end_column=i+1)
    ws.merge_cells(start_row=68, start_column=i, end_row=68, end_column=i+1)
    ws.merge_cells(start_row=81, start_column=i, end_row=81, end_column=i+1)
    ws.merge_cells(start_row=94, start_column=i, end_row=94, end_column=i+1)


for j in range (3, 108, 13):
    index = 0
    for k in range (3, 19, 2):
        ws.cell(row=j, column=k).value = target_list[index]
        ws.cell(row=j, column=k).alignment = Alignment(horizontal='center')
        ws.cell(row=j + 1, column=k).value = '平均'
        ws.cell(row=j + 1, column=k).alignment = Alignment(horizontal='center')
        ws.cell(row=j + 1, column=k + 1).value = '正負'
        ws.cell(row=j + 1, column=k + 1).alignment = Alignment(horizontal='center')
        index = index + 1

#read

source_index = 0
for source in source_list:

    wb_read = load_workbook(root_path + layer_path + excel_path + source + '_to_all.xlsx')
    pos_list = []
    pos_diff_list = []
    avg_list = []
    avg_diff_list = []
    for target in target_list:
        #target_index = 0
        ws_read = wb_read.get_sheet_by_name(source + '_' + target)
        pos = ws_read.cell(row=158, column=2).value
        #pos_percent = f"{pos:.0%}"
        pos_list.append(pos)
        #print(ws_read)
        #pos_list.append(ws_read.cell(row=158, column=2).value)
        #print(pos_list)
        pos_diff = ws_read.cell(row=159, column=2).value
        #pos_diff_percent = f"{pos_diff:.0%}"
        pos_diff_list.append(pos_diff)

        avg = ws_read.cell(row=164, column=2).value
        #avg_percent = f"{avg:.0%}"
        avg_list.append(avg)

        avg_diff = ws_read.cell(row=165, column=2).value
        #avg_diff_percent = f"{avg_diff:.0%}"
        avg_diff_list.append(avg_diff)

    target_index = 0
    for s in range (3, 19, 2):
        #s_count = 0
        ws.cell(row=(5 + (13 * source_index)) , column=s).value = pos_list[target_index]
        #print(pos_list[target_index])
        ws.cell(row=(5 + (13 * source_index)), column=s).alignment = Alignment(horizontal='center')
        ws.cell(row=(5 + (13 * source_index)), column=s+1).value = pos_diff_list[target_index]
        #print(pos_diff_list[target_index])
        ws.cell(row=(5 + (13 * source_index)), column=s+1).alignment = Alignment(horizontal='center')
        ws.cell(row=(6 + (13 * source_index)), column=s).value = avg_list[target_index]
        ws.cell(row=(6 + (13 * source_index)), column=s).alignment = Alignment(horizontal='center')
        ws.cell(row=(6 + (13 * source_index)), column=s+1).value = avg_diff_list[target_index]
        ws.cell(row=(6 + (13 * source_index)), column=s+1).alignment = Alignment(horizontal='center')

        target_index = target_index + 1
    source_index = source_index + 1












wb.save(root_path + layer_path + excel_path + "all_to_all.xlsx")



