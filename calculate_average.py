import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import numpy as np
import openpyxl

source_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']

source_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

target_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}


root_path = 'D:/ScottTsai/res_multi/auto_hyperparameter/'
layer_path = 'imp_v3_add3_with_kmm_dr_5Layer_auto_all/'
result_path = 'accuracy/result'
excel_path = 'accuracy/result/'


wb1 = load_workbook(root_path + layer_path + excel_path + 'ap_to_all.xlsx')
wb2 = load_workbook(root_path + layer_path + excel_path + 'ba_to_all.xlsx')
wb3 = load_workbook(root_path + layer_path + excel_path + 'ca_to_all.xlsx')
wb4 = load_workbook(root_path + layer_path + excel_path + 'ga_to_all.xlsx')
wb5 = load_workbook(root_path + layer_path + excel_path + 'mu_to_all.xlsx')
wb6 = load_workbook(root_path + layer_path + excel_path + 'ph_to_all.xlsx')
wb7 = load_workbook(root_path + layer_path + excel_path + 'pr_to_all.xlsx')
wb8 = load_workbook(root_path + layer_path + excel_path + 'to_to_all.xlsx')



ws1_list = wb1.worksheets
ws2_list = wb2.worksheets
ws3_list = wb3.worksheets
ws4_list = wb4.worksheets
ws5_list = wb5.worksheets
ws6_list = wb6.worksheets
ws7_list = wb7.worksheets
ws8_list = wb8.worksheets

if 'Sheet' in wb1.sheetnames:
    del wb1['Sheet']
    del wb2['Sheet']
    del wb3['Sheet']
    del wb4['Sheet']
    del wb5['Sheet']
    del wb6['Sheet']
    del wb7['Sheet']
    del wb8['Sheet']

wb1.save(root_path + layer_path + excel_path + 'ap_to_all.xlsx')
wb2.save(root_path + layer_path + excel_path + 'ba_to_all.xlsx')
wb3.save(root_path + layer_path + excel_path + 'ca_to_all.xlsx')
wb4.save(root_path + layer_path + excel_path + 'ga_to_all.xlsx')
wb5.save(root_path + layer_path + excel_path + 'mu_to_all.xlsx')
wb6.save(root_path + layer_path + excel_path + 'ph_to_all.xlsx')
wb7.save(root_path + layer_path + excel_path + 'pr_to_all.xlsx')
wb8.save(root_path + layer_path + excel_path + 'to_to_all.xlsx')





for source in source_list:
    target_df_file = pd.ExcelFile(root_path + layer_path + excel_path + source + '_to_all.xlsx')
    print(source)
    if source == 'ap':
        for sheet_num, sheets in enumerate (ws1_list):
            print(sheets)
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws1_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            #print(target_df.shape)
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            goal_ws.column_dimensions['A'].wth = 15
            goal_ws.cell(row=155, column= 1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            # Acc_positive
            for index in range (3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range (2, 6):
                goal_ws.cell (row = 155, column = s).value = acc_pos_list[s-2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            #pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            #print(pos_diff)
            goal_ws.cell(row=159, column=2).value = pos_diff

            #Acc_total
            for acc_index in range (0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range (2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            #max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            #total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff





    elif source == 'ba':
        for sheet_num, sheets in enumerate (ws2_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws2_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)
                #print(acc_pos_list)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]
                #print(acc_pos_list)

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff







    elif source == 'ca':
        for sheet_num, sheets in enumerate (ws3_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws3_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff

    elif source == 'ga':
        for sheet_num, sheets in enumerate (ws4_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws4_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff

    elif source == 'mu':
        for sheet_num, sheets in enumerate (ws5_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws5_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff

    elif source == 'ph':
        for sheet_num, sheets in enumerate (ws6_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws6_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff

    elif source == 'pr':
        for sheet_num, sheets in enumerate (ws7_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws7_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff

    else:
        for sheet_num, sheets in enumerate (ws8_list):
            acc_pos_list = []
            acc_total_list = []
            goal_ws = ws8_list[sheet_num]
            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.iloc[0:151]
            df_result_row = target_df.iloc[150]
            df_result_row = df_result_row.drop(0)

            target_df = pd.DataFrame(goal_ws.values)
            target_df = target_df.loc[0:150]
            target_df = target_df.dropna(axis=1)

            goal_ws.column_dimensions['A'].width = 15
            goal_ws.cell(row=155, column=1).value = 'Acc_positive'
            goal_ws.cell(row=155, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=156, column=1).value = 'max'
            goal_ws.cell(row=156, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=157, column=1).value = 'min'
            goal_ws.cell(row=157, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=158, column=1).value = 'Average'
            goal_ws.cell(row=158, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=159, column=1).value = '正負'
            goal_ws.cell(row=159, column=1).alignment = Alignment(horizontal='center')

            goal_ws.cell(row=161, column=1).value = 'Acc_total'
            goal_ws.cell(row=161, column=1).alignment = Alignment(horizontal = 'center')
            goal_ws.cell(row=162, column=1).value = 'max'
            goal_ws.cell(row=162, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=163, column=1).value = 'min'
            goal_ws.cell(row=163, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=164, column=1).value = 'Average'
            goal_ws.cell(row=164, column=1).alignment = Alignment(horizontal= 'center')
            goal_ws.cell(row=165, column=1).value = '正負'
            goal_ws.cell(row=165, column=1).alignment = Alignment(horizontal='center')

            for index in range(3, 23, 6):
                fn = df_result_row.iloc[index]
                tp = df_result_row.iloc[index + 1]
                acc_pos = tp / (tp + fn) * 100
                acc_pos = round(acc_pos, 3)
                acc_pos_list.append(acc_pos)

            for s in range(2, 6):
                goal_ws.cell(row=155, column=s).value = acc_pos_list[s - 2]

            # max, min
            max_pos = max(acc_pos_list)
            min_pos = min(acc_pos_list)
            goal_ws.cell(row=156, column=2).value = max_pos
            goal_ws.cell(row=157, column=2).value = min_pos

            # pos_avg
            pos_avg_list = acc_pos_list
            pos_avg_list.remove(max_pos)
            pos_avg_list.remove(min_pos)
            pos_avg = round(sum(pos_avg_list) / 2, 3)
            goal_ws.cell(row=158, column=2).value = pos_avg

            # +-
            pos_diff = max(pos_avg_list) - pos_avg
            goal_ws.cell(row=159, column=2).value = pos_diff

            # Acc_total
            for acc_index in range(0, 20, 6):
                result = df_result_row.iloc[acc_index]
                result = round(result, 3)
                acc_total_list.append(result)

            for t in range(2, 6):
                goal_ws.cell(row=161, column=t).value = acc_total_list[t - 2]

            # max, min
            max_total = max(acc_total_list)
            min_total = min(acc_total_list)
            goal_ws.cell(row=162, column=2).value = max_total
            goal_ws.cell(row=163, column=2).value = min_total

            # total_avg
            total_avg_list = acc_total_list
            total_avg_list.remove(max_total)
            total_avg_list.remove(min_total)
            total_avg = round(sum(total_avg_list) / 2, 3)
            goal_ws.cell(row=164, column=2).value = total_avg

            # +-
            total_diff = max(total_avg_list) - total_avg
            goal_ws.cell(row=165, column=2).value = total_diff





wb1.save(root_path + layer_path + excel_path + 'ap_to_all.xlsx')
wb2.save(root_path + layer_path + excel_path + 'ba_to_all.xlsx')
wb3.save(root_path + layer_path + excel_path + 'ca_to_all.xlsx')
wb4.save(root_path + layer_path + excel_path + 'ga_to_all.xlsx')
wb5.save(root_path + layer_path + excel_path + 'mu_to_all.xlsx')
wb6.save(root_path + layer_path + excel_path + 'ph_to_all.xlsx')
wb7.save(root_path + layer_path + excel_path + 'pr_to_all.xlsx')
wb8.save(root_path + layer_path + excel_path + 'to_to_all.xlsx')






