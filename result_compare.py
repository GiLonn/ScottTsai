import os
import pandas as pd
import numpy as np
import math
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

source_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']

source_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

target_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

root_path = 'D:/ScottTsai/res_multi/auto_hyperparameter/'
layer_path = 'imp_v3_add3_with_kmm_dr_5Layer_auto_all/'
result_path = 'accuracy/result'
excel_path = 'accuracy/result/'

goal_excel = load_workbook(root_path + layer_path + excel_path + "all_to_all.xlsx")
ws = goal_excel.get_sheet_by_name("1_to_2_average")
goal_excel.active = ws


#compare
for ROW in range(11, 104, 13):
    ws.cell(row = ROW, column = 2).value = 'Acc_Positive(Ours - Zhou)'
    ws.cell(row = ROW, column = 2).alignment = Alignment(horizontal='center')
    ws.cell(row = ROW + 1, column = 2).value = 'Acc_Total(Ours - Zhou)'
    ws.cell(row=ROW + 1, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row = ROW + 2, column = 2).value = '如果贏->1'
    ws.cell(row=ROW + 2, column=2).alignment = Alignment(horizontal='center')


for row_index_1 in range(5, 97, 13):
    row_index_2 = row_index_1 + 3
    for index in range(3, 58, 2):
        index_value1_pos = ws.cell(row = row_index_1, column = index).value
        index_value2_pos = ws.cell(row = row_index_2, column = index).value
        index_value1_total = ws.cell(row = row_index_1 + 1, column = index).value
        index_value2_total = ws.cell(row = row_index_2 + 1, column = index).value
        index_value1_diff_pos = ws.cell(row = row_index_1, column = index + 1).value
        index_value2_diff_pos = ws.cell(row = row_index_2, column = index + 1).value
        index_value1_diff_total = ws.cell(row=row_index_1 + 1, column=index + 1).value
        index_value2_diff_total = ws.cell(row=row_index_2 + 1, column=index + 1).value


        acc_positive = (index_value1_pos - (index_value2_pos * 100))
        acc_total = (index_value1_total - (index_value2_total*100))
        diff_positive = (index_value1_diff_pos - (index_value2_diff_pos * 100))
        diff_total = (index_value1_diff_total - (index_value2_diff_total * 100))
        acc_positive = round(acc_positive, 2)
        acc_total = round(acc_total, 2)
        diff_positive = round(diff_positive, 2)
        diff_total = round(diff_total, 2)



        ws.cell(row = row_index_1 + 6, column = index).value = acc_positive
        ws.cell(row = row_index_1 + 6, column = index).alignment = Alignment(horizontal='center')
        ws.cell(row = row_index_1 + 7, column = index).value = acc_total
        ws.cell(row=row_index_1 + 7, column=index).alignment = Alignment(horizontal='center')
        ws.cell(row = row_index_1 + 6, column = index + 1).value = diff_positive
        ws.cell(row=row_index_1 + 6, column=index + 1).alignment = Alignment(horizontal='center')
        ws.cell(row = row_index_1 + 7, column = index + 1).value = diff_total
        ws.cell(row=row_index_1 + 7, column=index + 1).alignment = Alignment(horizontal='center')

#winning combination
sum = 0
for b in range(12, 104, 13):
    for c in range(3, 58, 2):
        if ws.cell(row = b, column = c).value > 0:
            result = 1
            ws.cell(row = b + 1, column = c).value = 1
            ws.cell(row = b + 1, column = c).alignment = Alignment(horizontal='center')
            sum = sum + 1
        else:
            ws.cell(row=b + 1, column=c).value = 0
            ws.cell(row=b + 1, column=c).alignment = Alignment(horizontal='center')

ws.cell(row = 111, column = 2).value = 'winning combination'
ws.cell(row = 111, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 111, column = 3).value = sum
ws.cell(row = 111, column = 3).alignment = Alignment(horizontal='center')

#max and min
max_pos = 0
min_pos = 0
max_total = 0
min_total = 0
max_pos_diff = 0
min_pos_diff = 0
max_total_diff = 0
min_total_diff = 0

for row_index in range(11, 103, 13):
    for column_index in range(3, 58, 2):

        if ws.cell(row = row_index, column = column_index).value > max_pos:
            max_pos = ws.cell(row = row_index, column = column_index).value
        elif ws.cell(row = row_index, column = column_index).value < min_pos:
            min_pos = ws.cell(row = row_index, column = column_index).value

        if ws.cell(row = row_index + 1, column = column_index).value > max_total:
            max_total = ws.cell(row = row_index + 1, column = column_index).value

        elif ws.cell(row = row_index + 1, column = column_index).value < min_total:
            min_total = ws.cell(row = row_index + 1, column = column_index).value


        if ws.cell(row = row_index, column = column_index + 1).value > max_pos_diff:
            max_pos_diff = ws.cell(row = row_index, column = column_index + 1).value
        elif ws.cell(row = row_index, column = column_index + 1).value < min_pos_diff:
            min_pos_diff = ws.cell(row = row_index, column = column_index + 1).value

        if ws.cell(row = row_index + 1, column = column_index + 1).value > max_total_diff:
            max_total_diff = ws.cell(row = row_index + 1, column = column_index +1).value
        elif ws.cell(row = row_index + 1, column = column_index + 1).value < min_total_diff:
            min_total_diff = ws.cell(row = row_index + 1, column = column_index + 1).value

ws.cell(row = 112, column = 2).value = 'Max_acc_positive'
ws.cell(row = 112, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 112, column = 3).value = max_pos
ws.cell(row = 113, column = 2).value = 'Min_acc_positive'
ws.cell(row = 113, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 113, column = 3).value = min_pos
ws.cell(row = 114, column = 2).value = 'Max_acc_total'
ws.cell(row = 114, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 114, column = 3).value = max_total
ws.cell(row = 115, column = 2).value = 'Min_acc_total'
ws.cell(row = 115, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 115, column = 3).value = min_total

ws.cell(row = 116, column = 2).value = 'Max_diff_positive'
ws.cell(row = 116, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 116, column = 3).value = max_pos_diff
ws.cell(row = 117, column = 2).value = 'Min_diff_positive'
ws.cell(row = 117, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 117, column = 3).value = min_pos_diff
ws.cell(row = 118, column = 2).value = 'Max_diff_total'
ws.cell(row = 118, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 118, column = 3).value = max_total_diff
ws.cell(row = 119, column = 2).value = 'Min_diff_total'
ws.cell(row = 119, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 119, column = 3).value = min_total_diff


#average for each friuts
ws.cell(row = 121, column = 3).value = 'zhou'
ws.cell(row = 121, column = 3).alignment = Alignment(horizontal='center')
ws.cell(row = 121, column = 4).value = 'ours'
ws.cell(row = 121, column = 4).alignment = Alignment(horizontal='center')
for g in range(0, 8):
    ws.cell(row = 122 + g, column = 2).value = source_list[g]
    ws.cell(row = 122 + g, column = 2).alignment = Alignment(horizontal='center')
count_row = 0
for row_avg in range(6, 98, 13):

    sum_ours = 0
    sum_zhou = 0
    count = 0
    for index_avg in range (3, 58, 2):
        sum_ours = sum_ours + ws.cell(row = row_avg, column = index_avg).value
        sum_zhou = sum_zhou + (ws.cell(row = row_avg + 3, column = index_avg).value * 100)
        #print(sum_zhou)
        count = count + 1

    avg_ours = sum_ours / count
    avg_ours = round(avg_ours, 2)
    avg_zhou = sum_zhou / count
    avg_zhou = round(avg_zhou, 2)
    ws.cell(row = 122 + count_row, column = 3).value = avg_zhou
    ws.cell(row=122 + count_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=122 + count_row, column=4).value = avg_ours
    ws.cell(row=122 + count_row, column=4).alignment = Alignment(horizontal='center')

    count_row = count_row + 1

sum_1 = 0
sum_2 = 0
count = 0
for k in range(122, 130):
    sum_1 = sum_1 + ws.cell(row = k, column = 3).value
    sum_2 = sum_2 + ws.cell(row=k, column=4).value
    count = count + 1
avg_1 = sum_1 / count
avg_2 = sum_2 / count

avg1 = round(avg_1, 2)
avg2 = round(avg_2, 2)

ws.cell(row = 131, column = 2).value = 'total average'
ws.cell(row = 131, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 131, column = 3).value = avg1
ws.cell(row = 131, column = 3).alignment = Alignment(horizontal='center')
ws.cell(row = 131, column = 4).value = avg2
ws.cell(row = 131, column = 4).alignment = Alignment(horizontal='center')


best_index = []
worst_index = []
best_acc_ours = 0
worst_acc_ours = 100
for i in range(6, 98, 13):
    for j in range(3, 58, 2):
        if ws.cell(row = i, column = j).value > best_acc_ours:
            best_acc_ours = ws.cell(row = i, column = j).value
            best_index = [i, j]
        elif ws.cell(row = i, column = j).value < worst_acc_ours:
            worst_acc_ours = ws.cell(row = i, column = j).value
            worst_index = [i, j]

zhou_compare_best = ws.cell(row = best_index[0] + 3, column = best_index[1]).value * 100
zhou_compare_worst = ws.cell(row = worst_index[0] + 3, column = worst_index[1]).value * 100
zhou_compare_best = round(zhou_compare_best, 2)
zhou_compare_worst = round(zhou_compare_worst, 2)

#print(best_index)
#print(worst_index)

ws.cell(row = 134, column = 2).value = 'best'
ws.cell(row = 134, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 135, column = 2).value = 'worst'
ws.cell(row = 135, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 133, column = 2).value = 'ours'
ws.cell(row = 133, column = 2).alignment = Alignment(horizontal='center')
ws.cell(row = 133, column = 3).value = 'row'
ws.cell(row = 133, column = 3).alignment = Alignment(horizontal='center')
ws.cell(row = 133, column = 4).value = 'col'
ws.cell(row = 133, column = 4).alignment = Alignment(horizontal='center')
ws.cell(row = 133, column = 5).value = 'zhou_acc'
ws.cell(row = 133, column = 5).alignment = Alignment(horizontal='center')
ws.cell(row = 133, column = 6).value = 'ours_acc'
ws.cell(row = 133, column = 6).alignment = Alignment(horizontal='center')

ws.cell(row = 134, column = 3).value = best_index[0]
ws.cell(row = 134, column = 4).value = best_index[1]
ws.cell(row = 134, column = 5).value = zhou_compare_best
ws.cell(row = 134, column = 6).value = best_acc_ours

ws.cell(row = 135, column = 3).value = worst_index[0]
ws.cell(row = 135, column = 4).value = worst_index[1]
ws.cell(row = 135, column = 5).value = zhou_compare_worst
ws.cell(row = 135, column = 6).value = worst_acc_ours









goal_excel.save(root_path + layer_path + excel_path + "all_to_all.xlsx")
