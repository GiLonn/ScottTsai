import os
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
import openpyxl





root_path = 'D:/ScottTsai/res_multi/auto_hyperparameter/'
layer_path = 'imp_v3_add3_with_kmm_dr_5Layer_auto_all/'
result_path = 'accuracy/result'
excel_path = 'accuracy/result/'
#create folder and excel
create_folder_path = os.path.join(root_path, layer_path, result_path)
create_excel_path = os.path.join(root_path, layer_path, excel_path)

if os.path.exists(create_folder_path) == False:
    os.mkdir(create_folder_path)

wb1 = Workbook()
wb2 = Workbook()
wb3 = Workbook()
wb4 = Workbook()
wb5 = Workbook()
wb6 = Workbook()
wb7 = Workbook()
wb8 = Workbook()



wb1.save(create_excel_path + 'ap_to_all.xlsx')
wb2.save(create_excel_path + 'ba_to_all.xlsx')
wb3.save(create_excel_path + 'ca_to_all.xlsx')
wb4.save(create_excel_path + 'ga_to_all.xlsx')
wb5.save(create_excel_path + 'mu_to_all.xlsx')
wb6.save(create_excel_path + 'ph_to_all.xlsx')
wb7.save(create_excel_path + 'pr_to_all.xlsx')
wb8.save(create_excel_path + 'to_to_all.xlsx')

#read

source = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']

source_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
target_list = ['ap', 'ba', 'ca', 'ga', 'mu', 'ph', 'pr', 'to']
source_path_list = ["apple/", "banana/", "carambola/", "guava/", "muskmelon/", "peach/",
                    "pear/", "tomato/"]
to = 'to/'
lr = '0.0001/'

target_path_list = ["apple/", "banana/", "carambola/", "guava/", "muskmelon/", "peach/", "pear/", "tomato/"]

result_name = ['test_1.csv', 'test_2.csv', 'test_3.csv', 'test_4.csv']
read_path = root_path + layer_path + 'accuracy/'

data_all = pd.DataFrame()
data_space = pd.DataFrame({'space' : " "}, index = [0])
data_result = pd.DataFrame()

#save
save_path = root_path + layer_path + 'accuracy/' + 'result/'
source_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}

target_dict = {"apple" : "ap", "banana": "ba", "carambola" : "ca", "guava" : "ga", "muskmelon" : "mu",
               "peach" : "ph", "pear" : "pr", "tomato" : "to"}



for source in source_path_list:
    for target in target_path_list:
        source1 = source[:-1]
        target1 = target[:-1]
        save_excel_path = os.path.join(save_path, source_dict[source1] + '_to_' + 'all.xlsx')
        for result_time in range(0, 4):
            read_result = result_name[result_time]
            read_excel_path = os.path.join(read_path, source, to, target, lr, read_result)
            data = pd.read_csv(read_excel_path, names = [read_result, 'tn', 'fp', 'fn', 'tp'])
            data_all = pd.concat([data_all, data], axis=1)
            data_all = pd.concat([data_all, data_space], axis=1)
        data_result = data_all
        data_result.index = data_result.index + 1

# sum up all the result
        with pd.ExcelWriter(save_excel_path, mode= 'a', engine = 'openpyxl') as writer:
            book = load_workbook(save_excel_path)
            data_result.to_excel(writer, sheet_name=source_dict[source1] + '_' + target_dict[target1])
            writer.save()
        data_all = pd.DataFrame()






